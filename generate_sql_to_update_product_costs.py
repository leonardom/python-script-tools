import sys
import re
from openpyxl import load_workbook


def main(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    # print(sheet["A1"].value)

    sql = '''
      UPDATE estoque.produtos SET
        custo = :preco,
        custo_compra = :preco,
        customedio = :preco,
        preco = :preco + (:preco * lucro / 100),
        preco_minimo = (:preco + (:preco * lucro / 100)) - ((:preco + (:preco * lucro / 100)) * 0.3)
        classificacao_fiscal = :ncm
      WHERE codigo = ':produto';
    '''
    ncm, produto, preco = None, None, None

    for row in sheet.iter_rows():
      if row[0].value != None and row[0].value != 'NCM/SH':
        ncm = row[0].value

      if row[1].value != None:
        produto = row[1].value

      if row[3].value != None:
        preco = row[3].value

      if ncm != None and produto != None and preco != None:
          sql_cmd = sql.replace(':preco', str(preco))
          sql_cmd = sql_cmd.replace(':produto', produto)
          sql_cmd = sql_cmd.replace(':ncm', str(ncm))

          print(sql_cmd)
          ncm, produto, preco = None, None, None


if __name__ == "__main__":
    main(sys.argv[1])
