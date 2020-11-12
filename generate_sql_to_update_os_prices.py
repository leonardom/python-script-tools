import sys
import re
from openpyxl import load_workbook


def main(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    # print(sheet["A1"].value)

    sql = '''
      UPDATE manufatura.qualidade_ordens_servicos_itens itm SET
        preco_custo = :preco
      FROM manufatura.qualidade_ordens_servicos os
      WHERE os.id = itm.numero
        AND itm.codigo = ':produto'
        AND os.id = :os;

      UPDATE estoque.produtos SET
        custo = :preco,
        custo_compra = :preco,
        customedio = :preco
      WHERE codigo = ':produto';
    '''
    os, produto, preco = None, None, None

    for row in sheet.iter_rows():
        # for cell in row:
        #     print(cell.value, end=" ")
        if row[0].value != None and 'OS. ' in row[0].value:
            os = re.sub(pattern=r'(\w+)\.\s(\d+)',
                        repl='\\2', string=row[0].value)

        if row[0].value != None and 'MN' in row[0].value:
            produto = row[0].value

        if row[1].value != None and 'R$: ' in row[1].value:
            preco = re.sub(pattern=r'(\w+)\$:\s*(\d+),(\d+)\s*(\w+)',
                           repl='\\2.\\3', string=row[1].value)

        if os != None and produto != None and preco != None:
            sql_cmd = sql.replace(':preco', preco)
            sql_cmd = sql_cmd.replace(':produto', produto)
            sql_cmd = sql_cmd.replace(':os', os)

            print(sql_cmd)
            #print(os, produto, preco)
            produto, preco = None, None


if __name__ == "__main__":
    main(sys.argv[1])
